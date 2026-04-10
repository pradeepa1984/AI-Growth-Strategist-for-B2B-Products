"""
reader.py — Load the Excel workbook and convert each row into a normalised dict.

Skill-tag columns (binary Yes/None) are collapsed into a Python list called
`skills_from_sheet`, which is later merged with any skills scraped from LinkedIn.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

from config import (
    COL_ABOUT, COL_CITY, COL_COMPANY, COL_COUNTRY, COL_EMAIL,
    COL_FIRST_NAME, COL_FOLLOWERS, COL_GEO, COL_LAST_NAME,
    COL_POSITION, COL_STATE, COL_STATUS, COL_URL,
    INPUT_FILE, SKILL_TAG_COLUMNS,
)

logger = logging.getLogger(__name__)


def _cell(row_dict: dict[str, Any], key: str) -> str | None:
    """Return stripped string value or None."""
    val = row_dict.get(key)
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def load_leads(path: str | Path | None = None) -> list[dict]:
    """
    Read the Excel file and return a list of normalised lead dicts.

    Each dict contains:
        name, linkedin_url, email, title, company, location,
        skills_from_sheet, about, geo_exposure, status, followers
    """
    file_path = Path(path or INPUT_FILE)
    if not file_path.exists():
        # Try relative to script location
        file_path = Path(__file__).parent.parent / INPUT_FILE

    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    logger.info("Loading workbook: %s", file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Build header → column-index map
    headers: list[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else "")

    leads: list[dict] = []

    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        row: dict[str, Any] = dict(zip(headers, row_cells))

        first = _cell(row, COL_FIRST_NAME) or ""
        last  = _cell(row, COL_LAST_NAME)  or ""
        name  = f"{first} {last}".strip() or None

        url = _cell(row, COL_URL)
        if not url:
            logger.debug("Skipping row with no URL: %s", name)
            continue

        # Collapse skill-tag columns into a list
        skills_from_sheet: list[str] = [
            col for col in SKILL_TAG_COLUMNS
            if str(row.get(col, "") or "").strip().lower() == "yes"
        ]

        # Build location string
        parts = filter(None, [
            _cell(row, COL_CITY),
            _cell(row, COL_STATE),
            _cell(row, COL_COUNTRY),
        ])
        location = ", ".join(parts) or None

        leads.append({
            "name":              name,
            "linkedin_url":      url,
            "email":             _cell(row, COL_EMAIL),
            # These may be filled by scraper if blank
            "title":             _cell(row, COL_POSITION),
            "company":           _cell(row, COL_COMPANY),
            "location":          location,
            "skills_from_sheet": skills_from_sheet,
            "about":             _cell(row, COL_ABOUT),
            "geo_exposure":      _cell(row, COL_GEO),
            "status":            _cell(row, COL_STATUS),
            "followers":         row.get(COL_FOLLOWERS),
            # Filled downstream
            "skills":            [],
            "industry":          None,
            "experience_level":  None,
        })

    wb.close()
    logger.info("Loaded %d leads", len(leads))
    return leads
